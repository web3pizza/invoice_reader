import os
import io
import glob
from pdf2image import convert_from_path
import pandas as pd
from google.cloud import vision
from google.cloud.vision import types
from openpyxl import Workbook

# Load credentials from JSON key file
os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = "/Users/aaron/keys/new-key1.json"

# Initialize the client
client = vision.ImageAnnotatorClient()

# Set up a list of keywords to search for in the invoice text
keywords = ["invoice", "amount", "date", "product", "description"]

# Set up the directory for storing Excel files
if not os.path.exists("invoices"):
    os.mkdir("invoices")

# Set up the directory for storing processed PDF files
if not os.path.exists("processed"):
    os.mkdir("processed")

# Prompt user for path to scanned image
image_path = input("Enter the path to the scanned image: ")

# Convert the PDF file to images
images = convert_from_path(image_path)

# Loop through the images and detect text in each one
text = ""
for i, image in enumerate(images):
    with io.BytesIO() as output:
        image.save(output, format="JPEG")
        content = output.getvalue()

    # Detect text in the image
    image = types.Image(content=content)
    response = client.text_detection(image=image)
    texts = response.text_annotations

    # Concatenate the text from all images
    for text_annotation in texts:
        text += text_annotation.description.lower() + "\n"
print(text)

# Check if any of the keywords are present in the text
""" if any(keyword in text for keyword in keywords):
    # Create a Pandas DataFrame from the text
    data = {"text": text.split("\n")}
    df = pd.DataFrame(data)

    # Print the DataFrame
    with pd.option_context("display.max_rows", None, "display.max_columns", None):
        print(df)

    # Extract the relevant information from the DataFrame
    invoice_num = df[df["text"].str.contains("QUANTITY")]["text"].values[0]
    invoice_date = df[df["text"].str.contains("date")]["text"].values[0]
    customer = df[df["text"].str.contains("customer")]["text"].values[0]
    amount = df[df["text"].str.contains("amount")]["text"].values[0]

    print("Creating New Escel File...")
    # Create a new Excel file and save the extracted data
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Invoice Number"
    ws["B1"] = "Invoice Date"
    ws["C1"] = "Customer"
    ws["D1"] = "Amount"
    ws["A2"] = invoice_num
    ws["B2"] = invoice_date
    ws["C2"] = customer
    ws["D2"] = amount
    wb.save("invoices/{}.xlsx".format(os.path.basename(image_path)))
    print("Excel File Created")
else:
    print("No keywords found in scanned image text.")

# Prompt user for directory to search for PDF files
pdf_dir = input("Enter the path to the directory containing PDF files to process: ")

# Loop through all PDF files in the specified directory
for filename in glob.glob(os.path.join(pdf_dir, "*.pdf")):
    # Process each PDF file
    process_pdf(filename)


# Define function to process PDF files
def process_pdf(filename):
    # Convert the PDF file to images
    images = convert_from_path(filename)

    # Loop through the images and detect text in each one
    text = ""
    for i, image in enumerate(images):
        with io.BytesIO() as output:
            image.save(output, format="JPEG")
            content = output.getvalue()

        # Detect text in the image
        image = types.Image(content=content)
        response = client.text_detection(image=image)
        texts = response.text_annotations

        # Concatenate the text from all images
        for text_annotation in texts:
            text += text_annotation.description.lower() + "\n"

    # Check if any of the keywords are present in the text
    if any(keyword in text for keyword in keywords):
        # Create a Pandas DataFrame from the text
        data = {"text": text.split("\n")}
        df = pd.DataFrame(data)

        # Extract the relevant information from the DataFrame
        invoice_num = df[df["text"].str.contains("QUANTITY")]["text"].values[0]
        invoice_date = df[df["text"].str.contains("date")]["text"].values[0]
        customer = df[df["text"].str.contains("customer")]["text"].values[0]
        amount = df[df["text"].str.contains("amount")]["text"].values[0]

        # Create a new Excel file and save the extracted data
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Invoice Number"
        ws["B1"] = "Invoice Date"
        ws["C1"] = "Customer"
        ws["D1"] = "Amount"
        ws["A2"] = invoice_num
        ws["B2"] = invoice_date
        ws["C2"] = customer
        ws["D2"] = amount
        wb.save("invoices/{}.xlsx".format(os.path.basename(filename)))
    else:
        print("No keywords found in scanned image text.")
        print(df)
 """
